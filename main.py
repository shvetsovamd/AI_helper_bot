import telegram
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ApplicationBuilder
import os
import PyPDF2
from docx import Document
from transformers import pipeline
import logging
from rake_nltk import Rake
import nltk
from keybert import KeyBERT
from langdetect import detect
from datetime import date
import openpyxl
from openpyxl import Workbook

# nltk.download('all')

# Настройка логирования для отладки
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

TOKEN = "MY_TOKEN"
DOWNLOAD_DIR = "downloads"  # Директория для временного хранения файлов
MAX_MESSAGE_LENGTH = 4096
MAX_MODEL_TOKENS = 1024
rake_extractor = Rake()  # Инициализируем RAKE для ключевых слов
file_path = "MyTable.xlsx"
sheet_name = "Лист1"

# Создаем директорию для загрузок, если её нет
if not os.path.exists(DOWNLOAD_DIR):
    os.makedirs(DOWNLOAD_DIR)


def insert_row_to_sheet(spreadsheet_name: str, worksheet_name: str, row_data: list):
    try:
        # Если файл существует, загружаем его
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            # Иначе создаем новую книгу и лист
            workbook = Workbook()
            workbook.remove(workbook.active)  # Удаляем стандартный "Sheet"
            print(f"Файл не найден. Создан новый файл: {file_path}")

        # Выбираем нужный лист. Если его нет, создаем
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)
            print(f"Лист '{sheet_name}' не найден. Создан новый лист.")

        # Добавляем строку данных в конец таблицы
        worksheet.append(row_data)

        # Сохраняем изменения в файле
        workbook.save(file_path)
        print(f"Успешно добавлена строка: {row_data} в файл '{file_path}' -> '{sheet_name}'.")

    except Exception as e:
        print(f"Произошла ошибка при работе с Excel файлом: {e}")


def get_keywords_keybert(text, top_n=10, keyphrase_length=1):
    model = KeyBERT('sentence-transformers/LaBSE')

    # Extract keywords/keyphrases
    keywords = model.extract_keywords(
        text,
        keyphrase_ngram_range=(1, keyphrase_length),
        use_mmr=True,  # Используем MMR для разнообразия результатов
        top_n=top_n
    )

    # Результат — список кортежей (слово, оценка)
    return ' '.join([word for word, score in keywords])


def extract_keywords(text):
    # NLTK уже загружен
    rake_extractor.extract_keywords_from_text(text)
    # Получаем топ-10 ключевых фраз с их оценками
    keywords_with_scores = rake_extractor.get_ranked_phrases_with_scores()

    # Форматируем результат
    if not keywords_with_scores:
        return "Ключевые слова не найдены."

    formatted_keywords = []
    # Берем топ-5 самых релевантных фраз
    for score, phrase in keywords_with_scores[:5]:
        formatted_keywords.append(f"- {phrase.capitalize()} (релевантность: {score:.2f})")

    return "\n".join(formatted_keywords)


async def start_mess(update: telegram.Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        'Привет! Я бот для чтения файлов. Отправь мне текстовый файл (.pdf, .docx, .txt, .md).')


async def incorrect_mess(update: telegram.Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text('Я жду файл или команды /start.')


# Функция для отправки длинного текста частями
async def send_long_text(update, file_name, text_content):
    if not text_content.strip():
        await update.message.reply_text(f"Файл '{file_name}' пуст или не содержит читаемого текста.")
        return

    text_chunks = [text_content[i:i + MAX_MESSAGE_LENGTH] for i in range(0, len(text_content), MAX_MESSAGE_LENGTH)]

    for i, chunk in enumerate(text_chunks):
        header = f"Содержимое файла '{file_name}' (часть {i + 1}/{len(text_chunks)}):\n\n" if len(
            text_chunks) > 1 else f"Содержимое файла '{file_name}':\n\n"
        await update.message.reply_text(header + chunk)


# Функция-заглушка для чтения PDF
def read_pdf(filepath):
    text = ""
    try:
        with open(filepath, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text() or ""
    except Exception as e:
        return f"Ошибка чтения PDF: {e}"
    return text


# Функция-заглушка для чтения DOCX
def read_docx(filepath):
    text = ""
    try:
        doc = Document(filepath)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        return f"Ошибка чтения DOCX: {e}"
    return text


async def upload_file(update: telegram.Update, context: ContextTypes.DEFAULT_TYPE):
    """Основной обработчик полученных документов."""
    document = update.message.document
    username = update.message.from_user.username

    if not document:
        await update.message.reply_text("Пожалуйста, прикрепите файл как документ.")
        return

    file_name = document.file_name
    file_size = document.file_size
    file_type = document.mime_type
    file_subtype = file_name.split('.')[-1].lower()

    supported_subtypes = ['pdf', 'docx', 'txt', 'md']

    if file_subtype not in supported_subtypes:
        await update.message.reply_text(f'Тип файла "{file_subtype}" не поддерживается для чтения.')
        return

    await update.message.reply_text(
        f"Характеристики файла:\nтип - {file_type}\nимя - {file_name}\nразмер - {file_size} байт"
    )

    # Определяем путь сохранения временного файла
    file_path_on_disk = os.path.join(DOWNLOAD_DIR, file_name)

    try:
        # 1. Скачиваем файл на диск (потоково, безопасно для памяти)
        file_object = await document.get_file()
        # Используем download_to_drive для потоковой записи на диск
        await file_object.download_to_drive(file_path_on_disk)

        await update.message.reply_text(f"Файл успешно сохранен локально: {file_path_on_disk}")

        text_content = ""

        # 2. Обрабатываем файл в зависимости от его типа
        if file_subtype in ['txt', 'md']:
            # Читаем простые текстовые файлы построчно (безопасно для памяти)
            try:
                with open(file_path_on_disk, 'r', encoding='utf-8') as f:
                    text_content = f.read()
            except UnicodeDecodeError:
                with open(file_path_on_disk, 'r', encoding='windows-1251') as f:
                    text_content = f.read()
                await update.message.reply_text("Внимание: Файл декодирован как 'windows-1251'.")

        elif file_subtype == 'pdf':
            text_content = read_pdf(file_path_on_disk)

        elif file_subtype == 'docx':
            text_content = read_docx(file_path_on_disk)

        # 3. Отправляем содержимое файла частями
        await send_long_text(update, file_name, text_content)
        await update.message.reply_text(f'Язык текста\n{detect(text_content)}')
        words = get_keywords_keybert(text_content)
        await update.message.reply_text(f'Ключевые слова\n{words}')
        keywords = extract_keywords(text_content)
        await update.message.reply_text(f'Ключевые фразы\n{keywords}')

    except Exception as e:
        await update.message.reply_text(f"Произошла критическая ошибка при обработке файла: {e}")
    finally:
        new_record = [date.today(), username, file_name, keywords, words]
        insert_row_to_sheet(file_name, sheet_name, new_record)

        # 4. ОБЯЗАТЕЛЬНО удаляем временный файл после обработки
        if os.path.exists(file_path_on_disk):
            os.remove(file_path_on_disk)
            print(f"Временный файл {file_path_on_disk} удален.")


def main():
    application = Application.builder().token(TOKEN).build()

    application.add_handler(CommandHandler("start", start_mess))
    application.add_handler(MessageHandler(filters.Document.ALL, upload_file))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, incorrect_mess))

    print("Бот запущен и слушает обновления...")
    application.run_polling(poll_interval=3)
    print("Бот остановлен.")


if __name__ == '__main__':
    main()
